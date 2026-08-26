import csv
import io
import json
import time
from functools import lru_cache
from typing import Any

from fastapi import HTTPException, UploadFile
from sqlalchemy import and_, desc, func
from sqlalchemy.orm import Session

from backend_core.smart_responses.matcher import find_best_match
from backend_core.smart_responses.models import SmartResponseLog, SmartResponseTemplate
from backend_core.smart_responses.schemas import SmartResponseCreate, SmartResponseUpdate

CACHE_TTL_SECONDS = 60
_cache: dict[tuple[str, str], tuple[float, list[SmartResponseTemplate]]] = {}

DEFAULT_RESPONSES = [
    {
        "rule_name": "Arabic Greeting",
        "intent": "greeting",
        "keywords": ["السلام عليكم", "مرحبا", "أهلا", "صباح الخير", "مساء الخير"],
        "match_type": "keyword",
        "response_text": "السلام عليكم ورحمة الله وبركاته.\nمرحبًا بك في منصة HSAAI.\nكيف يمكنني مساعدتك اليوم؟",
        "priority": 1000,
    },
    {
        "rule_name": "Thanks",
        "intent": "thanks",
        "keywords": ["شكرا", "شكراً", "تسلم", "يعطيك العافية"],
        "match_type": "keyword",
        "response_text": "على الرحب والسعة.\nيسعدني مساعدتك في أي وقت.",
        "priority": 900,
    },
    {
        "rule_name": "Who Are You",
        "intent": "who_are_you",
        "keywords": ["من أنت", "عرف بنفسك", "ما هي HSAAI"],
        "match_type": "keyword",
        "response_text": "أنا HSAAI Enterprise Assistant، المساعد الذكي المؤسسي المصمم لمساعدتك في البحث والمعرفة والإجراءات المؤسسية وتحليل المعلومات.",
        "priority": 900,
    },
    {
        "rule_name": "Help Capabilities",
        "intent": "help",
        "keywords": ["مساعدة", "ماذا تستطيع أن تفعل", "ما خدماتك"],
        "match_type": "keyword",
        "response_text": "أستطيع مساعدتك في البحث المعرفي، تحليل الملفات، الإجابة عن الأسئلة، إنشاء التقارير، دعم الإجراءات المؤسسية، وتحليل البيانات.",
        "priority": 850,
    },
    {
        "rule_name": "Contact Support",
        "intent": "contact_support",
        "keywords": ["الدعم الفني", "واجهت مشكلة", "لا يعمل النظام"],
        "match_type": "keyword",
        "response_text": "يسعدنا مساعدتك.\nيرجى وصف المشكلة بالتفصيل وسيتم توجيه طلبك إلى فريق الدعم.",
        "priority": 850,
    },
    {
        "rule_name": "Goodbye",
        "intent": "goodbye",
        "keywords": ["مع السلامة", "إلى اللقاء", "شكرا لكم"],
        "match_type": "keyword",
        "response_text": "شكرًا لاستخدامك منصة HSAAI.\nنتمنى لك يومًا موفقًا.",
        "priority": 800,
    },
]


def clear_cache() -> None:
    _cache.clear()


def _keywords_json(keywords: list[str] | None) -> str:
    return json.dumps(keywords or [], ensure_ascii=False)


def _keywords_list(template: SmartResponseTemplate) -> list[str]:
    try:
        return json.loads(template.keywords_json or "[]")
    except Exception:
        return []


def to_dict(template: SmartResponseTemplate) -> dict[str, Any]:
    return {
        "id": template.id,
        "tenant_id": template.tenant_id,
        "workspace_id": template.workspace_id,
        "rule_name": template.rule_name,
        "intent": template.intent,
        "keywords": _keywords_list(template),
        "match_type": template.match_type,
        "regex_pattern": template.regex_pattern or "",
        "response_text": template.response_text,
        "priority": template.priority,
        "enabled": template.enabled,
        "language": template.language,
        "usage_count": template.usage_count,
        "success_count": template.success_count,
        "fallback_count": template.fallback_count,
        "created_by": template.created_by,
        "updated_by": template.updated_by,
        "created_at": template.created_at,
        "updated_at": template.updated_at,
    }


def seed_defaults(db: Session, tenant_id: str = "default", workspace_id: str = "default") -> None:
    exists = db.query(SmartResponseTemplate).filter_by(tenant_id=tenant_id, workspace_id=workspace_id).first()
    if exists:
        return
    for item in DEFAULT_RESPONSES:
        db.add(
            SmartResponseTemplate(
                tenant_id=tenant_id,
                workspace_id=workspace_id,
                created_by="system",
                updated_by="system",
                rule_name=item["rule_name"],
                intent=item["intent"],
                keywords_json=_keywords_json(item["keywords"]),
                match_type=item["match_type"],
                response_text=item["response_text"],
                priority=item["priority"],
                enabled=True,
                language="ar",
            )
        )
    db.commit()
    clear_cache()


def list_templates(db: Session, tenant_id: str, workspace_id: str | None = None, include_disabled: bool = True) -> list[SmartResponseTemplate]:
    query = db.query(SmartResponseTemplate).filter(SmartResponseTemplate.tenant_id == tenant_id)
    if workspace_id:
        query = query.filter(SmartResponseTemplate.workspace_id == workspace_id)
    if not include_disabled:
        query = query.filter(SmartResponseTemplate.enabled.is_(True))
    return query.order_by(desc(SmartResponseTemplate.priority), SmartResponseTemplate.intent, SmartResponseTemplate.rule_name).all()


def _active_templates(db: Session, tenant_id: str, workspace_id: str) -> list[SmartResponseTemplate]:
    key = (tenant_id, workspace_id)
    cached = _cache.get(key)
    if cached and time.time() - cached[0] < CACHE_TTL_SECONDS:
        return cached[1]
    templates = (
        db.query(SmartResponseTemplate)
        .filter(
            and_(
                SmartResponseTemplate.tenant_id == tenant_id,
                SmartResponseTemplate.workspace_id == workspace_id,
                SmartResponseTemplate.enabled.is_(True),
            )
        )
        .order_by(desc(SmartResponseTemplate.priority))
        .all()
    )
    _cache[key] = (time.time(), templates)
    return templates


def detect_response(db: Session, message: str, tenant_id: str, workspace_id: str, user_id: str = "anonymous") -> dict[str, Any]:
    seed_defaults(db, tenant_id=tenant_id, workspace_id=workspace_id)
    templates = _active_templates(db, tenant_id, workspace_id)
    match = find_best_match(message, templates)
    if not match:
        db.add(
            SmartResponseLog(
                tenant_id=tenant_id,
                workspace_id=workspace_id,
                user_id=user_id,
                message=message,
                response_source="llm",
                score=0,
            )
        )
        db.commit()
        return {"matched": False, "agent": "HSAAI Enterprise Assistant", "source": "llm", "score": 0.0}

    template = match.template
    template.usage_count = (template.usage_count or 0) + 1
    template.success_count = (template.success_count or 0) + 1
    db.add(
        SmartResponseLog(
            tenant_id=tenant_id,
            workspace_id=workspace_id,
            user_id=user_id,
            message=message,
            matched_rule_id=template.id,
            intent=template.intent,
            score=match.score,
            response_source="smart_response",
        )
    )
    db.commit()
    clear_cache()
    return {
        "matched": True,
        "agent": "HSAAI Enterprise Assistant",
        "source": "smart_response",
        "intent": template.intent,
        "score": match.score,
        "rule_id": template.id,
        "message": template.response_text,
        "response": template.response_text,
        "rag_found": 0,
        "workspace_id": workspace_id,
    }


def create_template(db: Session, payload: SmartResponseCreate, tenant_id: str, actor: str) -> SmartResponseTemplate:
    template = SmartResponseTemplate(
        tenant_id=tenant_id,
        workspace_id=payload.workspace_id,
        rule_name=payload.rule_name,
        intent=payload.intent,
        keywords_json=_keywords_json(payload.keywords),
        match_type=payload.match_type,
        regex_pattern=payload.regex_pattern or "",
        response_text=payload.response_text,
        priority=payload.priority,
        enabled=payload.enabled,
        language=payload.language,
        created_by=actor,
        updated_by=actor,
    )
    db.add(template)
    db.commit()
    db.refresh(template)
    clear_cache()
    return template


def get_template(db: Session, template_id: int, tenant_id: str) -> SmartResponseTemplate:
    template = db.query(SmartResponseTemplate).filter_by(id=template_id, tenant_id=tenant_id).first()
    if not template:
        raise HTTPException(status_code=404, detail="Smart response template not found")
    return template


def update_template(db: Session, template_id: int, payload: SmartResponseUpdate, tenant_id: str, actor: str) -> SmartResponseTemplate:
    template = get_template(db, template_id, tenant_id)
    data = payload.model_dump(exclude_unset=True)
    for field, value in data.items():
        if field == "keywords":
            template.keywords_json = _keywords_json(value)
        else:
            setattr(template, field, value)
    template.updated_by = actor
    db.commit()
    db.refresh(template)
    clear_cache()
    return template


def delete_template(db: Session, template_id: int, tenant_id: str) -> None:
    template = get_template(db, template_id, tenant_id)
    db.delete(template)
    db.commit()
    clear_cache()


def analytics(db: Session, tenant_id: str, workspace_id: str | None = None) -> dict[str, Any]:
    base = db.query(SmartResponseLog).filter(SmartResponseLog.tenant_id == tenant_id)
    templates = db.query(SmartResponseTemplate).filter(SmartResponseTemplate.tenant_id == tenant_id)
    if workspace_id:
        base = base.filter(SmartResponseLog.workspace_id == workspace_id)
        templates = templates.filter(SmartResponseTemplate.workspace_id == workspace_id)
    total = base.count()
    smart = base.filter(SmartResponseLog.response_source == "smart_response").count()
    llm = base.filter(SmartResponseLog.response_source == "llm").count()
    top_rules = (
        templates.order_by(desc(SmartResponseTemplate.usage_count)).limit(10).all()
    )
    top_intents = (
        db.query(SmartResponseLog.intent, func.count(SmartResponseLog.id).label("count"))
        .filter(SmartResponseLog.tenant_id == tenant_id, SmartResponseLog.intent != "")
        .group_by(SmartResponseLog.intent)
        .order_by(desc("count"))
        .limit(10)
        .all()
    )
    return {
        "total_requests": total,
        "smart_response_hits": smart,
        "llm_fallbacks": llm,
        "match_rate": round(smart / total, 4) if total else 0,
        "llm_fallback_rate": round(llm / total, 4) if total else 0,
        "top_rules": [{"id": r.id, "rule_name": r.rule_name, "intent": r.intent, "usage_count": r.usage_count} for r in top_rules],
        "top_intents": [{"intent": item[0], "count": item[1]} for item in top_intents],
    }


def export_rows(db: Session, tenant_id: str, workspace_id: str | None = None) -> list[dict[str, Any]]:
    return [to_dict(t) for t in list_templates(db, tenant_id, workspace_id)]


def import_items(db: Session, items: list[dict[str, Any]], tenant_id: str, actor: str) -> dict[str, Any]:
    imported = 0
    errors: list[str] = []
    for index, item in enumerate(items, start=1):
        try:
            payload = SmartResponseCreate(
                rule_name=item.get("rule_name") or item.get("name") or f"Imported Rule {index}",
                intent=item.get("intent") or "general",
                keywords=item.get("keywords") if isinstance(item.get("keywords"), list) else str(item.get("keywords", "")).split("|"),
                match_type=item.get("match_type") or "keyword",
                regex_pattern=item.get("regex_pattern") or "",
                response_text=item.get("response_text") or item.get("response") or "",
                priority=int(item.get("priority") or 100),
                enabled=bool(item.get("enabled", True)),
                language=item.get("language") or "ar",
                workspace_id=item.get("workspace_id") or "default",
            )
            create_template(db, payload, tenant_id, actor)
            imported += 1
        except Exception as exc:
            errors.append(f"row {index}: {exc}")
    return {"imported": imported, "skipped": len(errors), "errors": errors}


async def parse_upload(file: UploadFile, fmt: str) -> list[dict[str, Any]]:
    raw = await file.read()
    text = raw.decode("utf-8-sig", errors="ignore")
    if fmt == "json":
        data = json.loads(text)
        return data if isinstance(data, list) else data.get("items", [])
    if fmt == "csv":
        return list(csv.DictReader(io.StringIO(text)))
    if fmt == "excel":
        try:
            from openpyxl import load_workbook
        except Exception as exc:
            raise HTTPException(status_code=500, detail=f"openpyxl is required for Excel import: {exc}")
        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(h or "").strip() for h in rows[0]]
        return [dict(zip(headers, row)) for row in rows[1:]]
    raise HTTPException(status_code=400, detail="Unsupported import format")
